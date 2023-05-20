from django.shortcuts import render, redirect
from .models import LeaveApplication
from django.contrib import messages
from django.http import JsonResponse
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from django.http import HttpResponse


def leave_applications(request):
    if request.method == 'POST':
        employee_name = request.POST.get('employee_name')
        leave_type = request.POST.get('leave_type')
        leave_start_date = request.POST.get('leave_start_date')
        leave_end_date = request.POST.get('leave_end_date')
        reason_for_leave = request.POST.get('reason_for_leave')

        leave = LeaveApplication.objects.create(
            employee_name=employee_name,
            leave_type=leave_type,
            start_date=leave_start_date,
            end_date=leave_end_date,
            reason_for_leave=reason_for_leave
        )
        leave.save()
        messages.success(request, "Leave applied successfully")
        return redirect('index')
    else:
        messages.warning(request, "Sorry, something went wrong...")
        return render(request, 'index1.html')

def index(request):
    leaves = LeaveApplication.objects.all()
    return render(request, 'index1.html', {'leaves': leaves})

def get_details(request):
    if request.method == 'POST':
        detail = request.POST.get('id')
        data = LeaveApplication.objects.filter(id=detail).values()
        return JsonResponse(list(data), safe=False)
    else:
        return JsonResponse({'error': 'Invalid request method.'}, status=400)

def remove_entry(request):
    if request.method == 'POST':
        entry_id = request.POST.get('entry_id')
        try:
            entry = LeaveApplication.objects.get(id=entry_id)
            entry.delete()

            messages.success(request, "Entry removed successfully")
            return redirect('index')
        except LeaveApplication.DoesNotExist:

            messages.success(request, "Entry not found")
            return redirect('index')
    else:
        return messages.warning( 'Invalid request method.')


def export_to_excel(request):
    leaves = LeaveApplication.objects.all()

    # Create a new workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Write the headers
    headers = ['ID', 'Employee Name', 'Leave Type', 'Start Date', 'End Date', 'Reason for Leave']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f'{col_letter}1'] = header
        sheet[f'{col_letter}1'].font = Font(bold=True)

    # Write the data rows
    for row_num, leave in enumerate(leaves, 2):
        sheet[f'A{row_num}'] = leave.id
        sheet[f'B{row_num}'] = leave.employee_name
        sheet[f'C{row_num}'] = leave.leave_type
        sheet[f'D{row_num}'] = leave.start_date
        sheet[f'E{row_num}'] = leave.end_date
        sheet[f'F{row_num}'] = leave.reason_for_leave

    # Set the column widths
    for col_num in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_num)
        sheet.column_dimensions[col_letter].width = 15

    # Set the response headers for the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=leave_applications.xlsx'

    # Save the workbook to the response
    workbook.save(response)

    return response
